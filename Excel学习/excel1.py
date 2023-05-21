#-*- coding:utf-8 -*-

from openpyxl import Workbook, load_workbook

#创建文件
# wb = Workbook()  #创建一个Excel文件在RAM
# sheet = wb.active
#
# print(sheet.title)
# sheet.title = "大王的菇凉们"
#
# sheet["B9"] = "black girl"
# sheet["C9"] = "172, 48, 84"
# sheet.append(["Rachel","170","49"])
#
# wb.save("excel_test.xlsx")  #保存Excel文件

#打开已有文件
wb = load_workbook("dashuju.xlsx")
print(wb.sheetnames)
print(wb.get_sheet_names)
sheet = wb['Sheet1']
# print(sheet["D5"])
# print(sheet['D5'].value)

# for cell in sheet["D5:D10"]:   #指定列的切片数据
#     print(cell[0].value)

#行的循环
# for row in sheet:
#     #print(row)
#     for cell in row:
#         print(cell.value,end=' ')
#     print()

#指定行和列
# for row in sheet.iter_rows(min_row=4, max_row=15, max_col=5):
#     for cell in row:
#         print(cell.value,end=' ')
#     print()

#按列遍历
# for col in sheet.columns:
#     for cell in col:
#         print(cell.value,end=' ')
#     print()

#指定列遍历
for col in sheet.iter_cols(min_col=3, max_col=4, min_row=4, max_row=15):
    for cell in col:
        print(cell.value,end=' ')
    print()

#删除
# wb.remove(sheet)
# del wb[sheet]

#更改文件
sheet['D5'] = "金角大王"
wb.save("dashuju.xlsx")





