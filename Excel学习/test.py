from openpyxl import load_workbook

#打开已有文件
wb = load_workbook('池外设备总量-2021-1-5.xlsx')
#print(wb.sheetnames)
sheet =wb['网络设备']



print(sheet.max_row)
print(sheet.max_column)

ips= []
for cell in list(sheet.columns)[2][1:-1]:
    #print(cell.value)
    if cell.value != None:
        ips.append(cell.value)
print(ips)
# with open('test.txt','w') as tx:
#     tx.write('switch vsys vr_OA\n')
#     tx.write('security-policy\n')
#     for ip in ips:
#         tx.write('rule name xxx\n')
#         tx.write('source-zone untrust_OA\n')
#         tx.write('destination-zone trust_OA\n')
#         tx.write('source-address '+ ip + '\n')
#         tx.write('action deny\n')
#         tx.write('\n')












    


