import sys
import time
import re
from collections import defaultdict,OrderedDict
import struct
import socket
import os
import json
from ctypes import CDLL,Structure,POINTER,c_char,c_char_p,c_int,c_ulong,c_bool,c_uint
from random import randint
testMatch = re.compile(r'\s*object-group ([\w\W]+)\n$')
objetcMatch = re.compile(r'object-group (?:ip address (?P<ipGroupName>[\d\w\W]+)|service (?P<portGroupName>[\d\w\W]+))\n?')
netMatch = re.compile(r'\s*\d+ network '
                      r'(?:host (?:address|name) (?P<host>[\w\W]+)|'
                      r'subnet (?P<sub>[\w\W]+?) (?:(?P<wd>wildcard)\s+)?(?P<mask>[\w\W]+)|'
                      r'range (?P<start>[\w\W]+) (?P<end>[\w\W]+)|'
                      r'exclude (?P<ex>[\w\W]+)|'
                      r'group-object (?P<gname>[\w\W]+)'
                      r')\s*\n?')
zoneMatch = re.compile(r'\s*security-zone name (?P<zonename>[\w\W]+)\s*\n?|\s*import interface (?P<int>[\w\W-]+)\s*\n?')
portMatch = re.compile(
    r'\s*\d+ service (?:(?P<protocal>(?:tcp|udp))(?: destination (?:(?P<symbol>eq|gt|lt|gte|lte) (?P<singlenum>\d+)|range (?P<start>\d+) (?P<end>\d+)))?|'
    r'group-object (?P<gname>[\w\W]+)|(?P<isicmp>icmp)'
    r')\s*\n?')
infoMatch = re.compile(r'\s*(:?description (?P<desc>[\w\W]+?)|'
                       r'action (?P<action>[\w\W]+)|'
                       r'source-zone (?P<fromzone>[\w\W]+)|'
                       r'destination-zone (?P<tozone>[\w\W]+)|'
                       r'source-ip(?: (?P<squote>[\w\W]+)|-host (?P<shost>[\w\W]+)|-subnet (?P<ssub>[\w\W]+) (?P<smask>[\w\W]+)|'
                       r'-range (?P<sstart>[\w\W]+) (?P<send>[\w\W]+))|'
                       r'destination-ip(?: (?P<dquote>[\w\W]+)|-host (?P<dhost>[\w\W]+)|-subnet (?P<dsub>[\w\W]+) (?P<dmask>[\w\W]+)'
                       r'-range (?P<dstart>[\w\W]+) (?P<dend>[\w\W]+))|'
                       r'service (?P<pquote>[\d\w\W-]+))\s*\n?')
                       #r'service (?:(?P<pn>\d+)|(?P<pquote>[\w\W-]+)))\s*\n?')
postcommandMatch=re.compile(r'export\s*([\w\W]+)|show\s*(rule|address|service|standardnet|standardport|sn|sp|s|r|a)\s*([\w\W]+)?\s*')
commandMatch = re.compile(r'(?P<isinit>init)\s*(?P<path>[\w\W]+)?|'
                          r'select (?:(?P<alldb>\s*\?)|(?P<db>[\w\W]+))|'
                          r'check\s+(?P<isshort>-s)\s+(?P<policy>[\w\W\"]+)|'
                          r'(?P<spolicy>[\w\W]+)')
addressMatch = re.compile(
    r'\"?(?P<bsub>[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3})\s*(?:(?P<ifwd>(?:w|wildcard|wc))\s+)?\s*(?P<bmask>[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3})\"?|'
    r'\"?(?P<start>[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3})[-~](?P<end>[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3})\"?|'
    r'(?P<asub>[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3})(?:/(?P<amask>[0-9]{1,2}))?')
szMatch = re.compile(r'\s*security-zone ([\w\W]+)')
ruMatch = re.compile(r'\s*rule \d+ name (?P<rulename>[\w\W]+)\s*\n?')
portAna = re.compile(r'(?P<sympol>(?:lt|lte|gt|gte|eq|\<|\<=|\>|\>=|=))?'
                     r'(?P<pro>tcp|udp|icmp)(?P<start>\d+)?(?:-(?P<end>\d+))?')
sympol = ['j', 'k', 'm', 'n', 'a', 'b', 'c', 'd']
char2sympol = {'gt': '>', 'gte': '>=', 'lt': '<', 'lte': '<=', 'eq': ''}
"""
objetcMatch用于匹配object-group 开头的语句，分为ip address和service两种，并提取出对应信息
netMatch用于匹配object-group ip address xxx下的所有语句，并提取出对应信息
zoneMatch用于提取security-zone信息
portMatch用于提取object-group service xxx下所有语句的对应信息
infoMatch用于提取rule语句中的所有有效信息
postcommandMatch用于交互式命令匹配(可忽略)
commandMatch用于交互式命令匹配(可忽略)
addressMatch用于专门提取ip地址信息
szMatch用于提取object-group里的secure-zone信息
ruMatch用于提取rule语句的rulename
portAna用于专门提取端口信息
sympol为随机变量符号，下面wildcard转换会用到
char2sympol为符号转化，端口转换和匹配会用到
"""

class standardNet(Structure):
    """
    标准网段,可以直接传入C
    用这个类代表C中有同样数据结构的结构体，从而可以用这个类来解析C返回的结构体指针,然后从内存中拿到对应的值

    _fields_ = [("base", c_ulong),  #注意因为c的结构体的对其问题，必须自行将field排好序，不然会发生数据错位
                ("offset", c_ulong),#c中的代码会自行结构体对齐，所以即使定义的类和结构体顺序一样，也可能错位
                ("itype", c_char)]
    """
    _fields_ = [("itype", c_char),
                ("base", c_ulong),
                ("offset", c_ulong)]

class standardPort(Structure):
    """
    标准端口，可以直接传入C
    用这个类代表C中有同样数据结构的结构体，从而可以用这个类来解析C返回的结构体指针,然后从内存中拿到对应的值
    """
    _fields_ = [("base", c_uint),
                ("offset",c_uint),
                ("protocal", c_char_p)] #t for tcp,u for udp,i for icmp

class structToPython(Structure):
    """
    调用C后的返回值
    用这个类代表C中有同样数据结构的结构体，从而可以用这个类来解析C返回的结构体指针,然后从内存中拿到对应的值
    """
    _fields_ = [("base", c_char_p),
                ("offset", c_char_p),
                ("itype", c_char)]


def sendToCGroup(source,targets,dll,isnum):
    """
    输入4个参数，source为一个网段三元组，targets为一个网段三元组列表，dll为加载的dll文件,isnum为三元组是否为数字
    isnum=false: source=('192.168.1.1','255.255.255.255','c') target=[('192.168.1.1','255.255.255.255','c'),(),()...]
    isnum=true:  source=(422609408,1014015,'w') target=[(423790448,4294967295,'c'),(),()...]
    网段三元组为开发时定义的数据结构，结构如下：
                            (base,offset,flag)
    当flag='c'时，代表这是一个普通网段掩码，base为网段，offset为掩码
    当flag='w'时，代表这是一个wildcard网段掩码，base为网段，offset为反掩码
    当flag='r'时，代表这是一个地址范围，base为起始地址，offset为结束地址

    返回时，如果source在targets任意一个网段里，返回时会返回一个匹配到的网段三元组，如果不在，则会返回None
    注:返回的网段三元组不一定是之前传入的targets中的一个，里面会做汇总，所以返回的的值是一个汇总后的匹配到的网段
    """
    dll.ifNetInTarget.restype = POINTER(structToPython)  #定义dll中函数的的返回类型(一个c指针)
    dll.ifNetInTargetLong.restype = POINTER(structToPython)
    length= len(targets)+1
    if isnum:
        addr = (c_ulong* length)()  #构造c结构的数组,以便传入dll中进行计算
        mask = (c_ulong* length)()
        addr[0]=source[0]
        mask[0]=source[1]
    else:
        addr=(c_char_p * length)()
        mask=(c_char_p * length)()
        addr[0] = bytes(source[0],'ascii')  #bytes化字符串
        mask[0] = bytes(source[1],'ascii')
    itype = (c_char * length)()
    itype[0]=bytes(source[2],'ascii')
    start=1
    for it in targets:
        if isnum:
            addr[start] = it[0]
            mask[start] = it[1]
        else:
            addr[start] = bytes(it[0],'ascii')
            mask[start] = bytes(it[1],'ascii')
        itype[start] = bytes(it[2],'ascii')
        start+=1
    nl=c_int(length)
    if isnum:
        res=dll.ifNetInTargetLong(addr,mask,itype,nl) #如果isnum为True,则将参数传入ifNetInTargetLong函数，返回的res为一个指针
    else:
        res=dll.ifNetInTarget(addr,mask,itype,nl) #如果isnum为False,则将参数传入ifNetInTarget函数
    base=res.contents.base #从指针中取出对应的值
    offset=res.contents.offset
    itype=res.contents.itype
    dll.recycle_res()  #回收内存，防止内存泄漏
    if itype==b'n':  #itype为'n',则未匹配到，返回None
        return None
    return (base,offset,itype)

def sendToCStandardIP(source,targets,dll):
    """
    功能同上一个函数，只是不同参数的实现
    :param source: 源standardnet
    :param targets: 目标standard集合
    :param dll: dll对象
    :return: 源是否在目标中
    """
    dll.ifNetInTargetStruct.restype = POINTER(structToPython)
    length = len(targets)
    tlist=(standardNet*length)()
    for i in range(length-1):
        tlist[i]=targets[i]
    length=c_int(length+1)
    res=dll.ifNetInTargetStruct(source,tlist,length)
    base = res.contents.base  # 从指针中取出对应的值
    offset = res.contents.offset
    itype = res.contents.itype
    dll.recycle_res()  # 回收内存，防止内存泄漏
    if itype == b'n':  # itype为'n',则未匹配到，返回None
        return None
    return (base, offset, itype)

class Firewalloperation(object):
    """
    防火墙基础类
    """
    def __init__(self):
        pass

    @staticmethod
    def ip2num(ip):
        """
        点分十进制ip转化成唯一的数字
        """
        return struct.unpack("!L", socket.inet_aton(ip))[0]

    @staticmethod
    def num2ip(num):
        """
        数字转换成点分十进制ip
        """
        a, c, b, d = (num & 0xff000000) >> 24, (num & 0x0000ff00) >> 8, (num & 0x00ff0000) >> 16, (
                num & 0x000000ff) >> 0
        return "{}.{}.{}.{}".format(a, b, c, d)

    @classmethod
    def commonTrans(cls, net, mask):
        """
        192.168.0.0 255.255.255.0这样的ip转换成192.168.1.0/24这样的精简形式
        """
        net = cls.ip2num(net)
        mask = cls.ip2num(mask)
        base = cls.num2ip(net & mask)
        fmask = 0
        while mask:
            mask = mask & (mask - 1)
            fmask += 1
        return "{}/{}".format(base, fmask)

    @staticmethod
    def portTrans(base, offset, sympol):
        """
        (2000,0,>)这样的输入转换成 (2001,65535)这样的输出
        """
        base, offset = int(base), int(offset)
        if sympol == "gt" or sympol == ">":
            base = base + 1
            offset = 65535
        elif sympol == "gte" or sympol == ">=":
            offset = 65535
        elif sympol == "lt" or sympol == "<":
            offset = base - 1
            base = 0
        elif sympol == "lte" or sympol == "<=":
            offset = base
            base = 0
        elif sympol == "eq" or sympol == "=":
            offset = base
        return base, offset

    @staticmethod
    def createStandardNet(base,offset,itype):
        """
        创建一个标准网段并返回，这个标准网段可以直接传入C
        """
        t = standardNet()
        base = H3CFwoperation.ip2num(base)
        offset = H3CFwoperation.ip2num(offset)
        if itype == 'c':
            base =c_ulong((base & 4294967295) & offset)
        elif itype == 'w':
            base =c_ulong(base &reversemask(offset))
        t.itype = bytes(itype, 'ascii')
        t.base= base
        t.offset = c_ulong(offset&4294967295)
        return t

    @staticmethod
    def createStandardPort(base, offset, protocal):
        """
        创建一个标准端口并返回，这个标准端口可以直接传入C
        """
        t = standardPort()
        t.base = c_uint(base)
        t.offset = c_uint(offset)
        t.protocal = bytes(protocal, 'ascii')
        return t

    @staticmethod
    def randIpPair(s1=0,e1=4294967295,s2=0,e2=4294967295):
        """
        生成一对随机ip，随机范围分别由s1,e1和s2,e2控制
        """
        a=randint(s1,e1)
        b=randint(s2,e2)
        return Firewalloperation.num2ip(a),Firewalloperation.num2ip(b)

reversemask = lambda x: (~(int(x) & 4294967295)) & 4294967295  #反掩码函数

class H3CFwoperation(Firewalloperation):
    """
    H3C防火墙操作类
    """
    def __init__(self):
        """
        net2standard->key:网段(字符串) value:标准网段
        short2mask->key:短掩码 value:长掩码
        port2standard->key:端口(字符串) value:标准端口
        addrQuote和portQuote都为用于引用计数的字典
        zone2interface->key:zone名 value:交换机接口名
        """
        super(H3CFwoperation, self).__init__()
        self.net2standard={}
        self.short2mask = {}
        self.port2standard={}
        self.addrQuote = defaultdict(int)
        self.portQuote = defaultdict(int)
        self.zone2interface =None
        self.zonecount = None
        for i in range(0, 32):
            self.short2mask[str(32 - i)] = (4294967295 >> i) << i

    @property
    def rp(self):
        """
        生成随机ip对
        """
        start, end = self.randIpPair(0, 4294967295, 0, 4294967295)
        return "{},{}".format(start, end)

    def rpc(self,s1=0,e1=4294967295,s2=0,e2=4294967295):
        """
        根据范围s1,e1生成第一个ip,s2,e2生成第二个ip
        """
        start,end=self.randIpPair(s1,e1,s2,e2)
        return "{},{}".format(start,end)

    @property
    def function(self):
        """
        打印出部分可使用的类信息
        """
        msg ="  ip2num        用于将ip转化成数字,使用如下:ip2num('192.168.1.1')\n"
        msg+="  num2ip        用于将数字转化成ip,使用如下:num2ip(20000)\n"
        msg+="  infoSummary   输出当前的h3c防火墙详细信息,输入infoSummary即可\n"
        msg+="  wildCardTrans 用于wildcard转化,使用如下:wildCardTrans('10.0.24.0','0.255.0.255')\n"
        msg+="  rp(c)         用于输出随机ip对,使用如下:1.rp 2.rpc(0,100,1000,2000)\n"
        msg+="  exportExcel   用于导出excel,已被export覆盖,原始使用如下:exportExcel('随机名字')\n"
        msg+="  (某属性)       直接访问内部属性，如net2standard/short2mask/port2standard/zone2interface"
        return msg

    @staticmethod
    def wildCardTrans(net, rmask):
        """
        用于把使用wildcard的ip转化成单一的字符串
        """
        za = net.split(".")
        ka = rmask.split(".")
        reversemask = lambda x: (~(int(x) & 255)) & 255
        fullIp = []
        for i in range(0, 4):
            mask = reversemask(ka[i])
            base = int(za[i]) & mask
            last, factor, sysp, final = 1, 1, 'p', []
            if base:
                final.append(str(base))
            for j in range(0, 8):
                t = mask & 0x1
                if t == 0:
                    if t != last:
                        sysp = sympol[j]
                        base = 1 << j
                        factor = 1
                    else:
                        factor += 1
                else:
                    if t != last:
                        final.append("{}*{}<{}~{}>".format(base, sysp, 0, (1 << factor) - 1))
                last = t
                mask = mask >> 1
            if t == 0:
                final.append("{}*{}<{}~{}>".format(base, sysp, 0, (1 << factor) - 1))
            fullIp.append("[{}]".format(' + '.join(final)))
        return '.'.join(fullIp)

    def createStandardNet2dict(self,key,base,offset,itype):
        """
        创建标准网段，并将其放入本类的net2standard字典中
        """
        if key not in self.net2standard:
            t=self.createStandardNet(base,offset,itype)
            self.net2standard[key]=t

    def createStandardPort2dict(self,key,base,offset,protocal,sympol=""):
        """
        创建标准端口，并将其放入本类的port2standard字典中
        """
        base,offset=self.portTrans(base,offset,sympol)
        if key not in self.port2standard:
            t=self.createStandardPort(base,offset,protocal)
            self.port2standard[key]=t

    def configTrans(self, path):
        """
        读入配置文件，并从中提取出有效的信息(rule,address,service)
        """

        with open(path, 'rb') as f:
            """
            with open下的语句为逐行读入提供的文本文件，并用预设好的正则去获取相应的信息，将之存入对应的字典
            """
            ls = f.readlines()
            readAddr, readPort, addrName, portName, ruleName, securityzone, nowzone = 0, 0, '', '', '', '', ''
            obj2ipaddr, obj2port, rule2all, zone2interface = defaultdict(dict), defaultdict(dict), OrderedDict(), defaultdict(list)
            direRead = ['action', 'fromzone', 'tozone', 'squote', 'dquote', 'pquote', 'desc']
            for l in ls:
                try: #尝试使用utf-8解码
                    l = l.decode("utf-8").strip()
                except: #utf-8失败再使用gbk解码
                    l = l.decode("gbk").strip()
                tm = zoneMatch.match(l)
                if tm:
                    tg = tm.groupdict()
                    if tg['zonename']:
                        nowzone = tg['zonename']
                    if nowzone and tg['int']:
                        zone2interface[nowzone].append(tg['int'])
                if "object-group" in l:
                    st = objetcMatch.match(l)
                    gst = st.groupdict()
                    if gst['ipGroupName']:
                        addrName = gst['ipGroupName']
                        obj2ipaddr[addrName] = {"positive": [], "negative": [], "quote": [], "szone": ""}
                    elif gst['portGroupName']:
                        portName = gst['portGroupName']
                        obj2port[portName] = {"port": [], "quote": [] }
                    continue
                elif "#" in l:
                    addrName = portName = ruleName = ""
                if addrName:
                    if "security-zone" in l:
                        securityzone = szMatch.findall(l)[0]
                        obj2ipaddr[addrName]["szone"] = securityzone.strip("\n")
                    elif 'network' in l:
                        res = netMatch.match(l)
                        if res:
                            gd = res.groupdict()
                            if gd['host']:
                                obj2ipaddr[addrName]["positive"].append("{}/32".format(gd.get('host')))
                                self.createStandardNet2dict("{}/32".format(gd.get('host')),gd.get('host'),'255.255.255.255', 'c')
                            elif gd['sub']:
                                if gd.get('wd'):
                                    t = self.wildCardTrans(gd.get('sub'), gd.get('mask'))
                                    obj2ipaddr[addrName]["positive"].append(t)
                                    self.createStandardNet2dict(t,gd.get('sub'),gd.get('mask'),'w')
                                else:
                                    t = self.commonTrans(gd.get('sub'), gd.get('mask'))
                                    obj2ipaddr[addrName]["positive"].append(t)
                                    self.createStandardNet2dict(t,gd.get('sub'),gd.get('mask'),'c')
                            elif gd['start']:
                                t = "{}~{}".format(gd.get('start'), gd.get('end'))
                                obj2ipaddr[addrName]["positive"].append(t)
                                self.createStandardNet2dict(t, gd.get('start'), gd.get('end'), 'r')
                            elif gd['ex']:
                                obj2ipaddr[addrName]["negative"].append("{}/32".format(gd.get('ex')))
                                self.createStandardNet2dict("{}/32".format(gd.get('ex')),gd.get('ex'),'255.255.255.255', 'c')
                            elif gd['gname']:
                                obj2ipaddr[addrName]["quote"].append(gd.get('gname'))
                elif portName:
                    res = portMatch.match(l)
                    if res:
                        gd = res.groupdict()
                        if gd['symbol']:
                            t="{}{}{}".format(char2sympol[gd['symbol']],gd.get('protocal'), gd.get('singlenum'))
                            obj2port[portName]['port'].append(t)
                            self.createStandardPort2dict(t,gd.get('singlenum'),0,gd.get('protocal'),gd['symbol'])
                        elif gd['start']:
                            t="{}{}-{}".format(gd.get('protocal'),gd['start'],gd.get('end'))
                            obj2port[portName]['port'].append(t)
                            self.createStandardPort2dict(t, gd['start'],gd.get('end'),gd.get('protocal'))
                        elif gd['gname']:
                            obj2port[portName]['quote'].append(gd['gname'])
                        elif gd['isicmp']:
                            obj2port[portName]['port'].append('icmp')
                            self.createStandardPort2dict(t, 0, 0, 'icmp')
                if "rule" in l:
                    rs = ruMatch.match(l)
                    if rs:
                        grs = rs.groupdict()
                        if grs['rulename']:
                            ruleName = grs['rulename']
                            rule2all[ruleName] = {'action': 'deny', "fromzone": '', "tozone": '', 'spositive': [],
                                                  "dpositive": [], 'snegative': [], "dnegative": [], "squote": [],
                                                  "dquote": [], 'pquote': [],
                                                  "service": {"port": [],'common': []},
                                                  "desc": ""}
                if ruleName:
                    res = infoMatch.match(l)
                    # print(l)
                    if res:
                        gd = res.groupdict()
                        for item in direRead:
                            if gd[item]:
                                if isinstance(rule2all[ruleName][item], list):
                                    rule2all[ruleName][item].append(gd[item])
                                else:
                                    rule2all[ruleName][item] = gd[item]
                        if gd['shost']:
                            rule2all[ruleName]["spositive"].append("{}/32".format(gd['shost']))
                            self.createStandardNet2dict("{}/32".format(gd['shost']),gd['shost'], '255.255.255.255', 'c')
                        elif gd['ssub']:
                            t = self.commonTrans(gd.get('ssub'), gd.get('smask'))
                            rule2all[ruleName]["spositive"].append(t)
                            self.createStandardNet2dict(t, gd.get('ssub'), gd.get('smask'), 'c')
                        elif gd['sstart']:
                            t = "{}~{}".format(gd.get('sstart'), gd.get('send'))
                            rule2all[ruleName]["spositive"].append(t)
                            self.createStandardNet2dict(t, gd.get('sstart'), gd.get('send'), 'r')
                        elif gd['dhost']:
                            rule2all[ruleName]["dpositive"].append("{}/32".format(gd['dhost']))
                            self.createStandardNet2dict("{}/32".format(gd['dhost']), gd['dhost'], '255.255.255.255','c')
                        elif gd['dsub']:
                            t = self.commonTrans(gd.get('dsub'), gd.get('dmask'))
                            rule2all[ruleName]["dpositive"].append(t)
                            self.createStandardNet2dict(t, gd.get('sub'), gd.get('mask'), 'c')
                        elif gd['dstart']:
                            t = "{}~{}".format(gd.get('dstart'), gd.get('dend'))
                            rule2all[ruleName]["dpositive"].append(t)
                            self.createStandardNet2dict(t, gd.get('dstart'), gd.get('dend'), 'c')
                        """elif gd['pn']:
                            rule2all[ruleName]['service']['port'].append("tcp{}".format(gd['pn']))
                            rule2all[ruleName]['service']['port'].append("udp{}".format(gd['pn']))
                            self.createStandardPort2dict("tcp{}".format(gd['pn']), gd.get('pn'), gd.get('pn'), 'tcp','')
                            self.createStandardPort2dict("udp{}".format(gd['pn']), gd.get('pn'), gd.get('pn'), 'udp','')"""
        """
        配置的一次性解析已经完成，接下来要对嵌套的object-group进行遍历，然后添加到rule2all,obj2port,obj2ipaddr字典中
        """
        self.zone2interface = zone2interface

        def traveladdr(obj, objd):
            """
            内部函数，用于遍历address
            """
            positve, negative = [], []
            if not objd[obj]:
                return None
            else:
                self.addrQuote[obj] += 1
                positve += objd[obj]['positive']
                negative += objd[obj]['negative']
                for item in objd[obj]['quote']:
                    out = traveladdr(item, objd)
                    if out:
                        positve += out[0]
                        negative += out[1]
                return list(set(positve)), list(set(negative))

        def travelport(obj, objd):
            """
            内部函数，用于遍历port
            """
            port,common=[],[]
            if not objd[obj]:
                return None
            else:
                self.portQuote[obj] += 1
                port+=objd[obj]['port']
                for item in objd[obj]['quote']:
                    out= travelport(item, objd)
                    if out:
                        port += out[0]
                        common += out[1]
                    else:
                        common.append(item)
                return list(set(port)),list(set(common))

        for ruleName in rule2all:
            squote = rule2all[ruleName]['squote']
            dquote = rule2all[ruleName]['dquote']
            pquote = rule2all[ruleName]['pquote']
            for key, quote in zip(['s', 'd'], [squote, dquote]):
                if quote:
                    for q in quote:
                        res = traveladdr(q, obj2ipaddr)
                        if res:
                            positive, negative = traveladdr(q, obj2ipaddr)
                            rule2all[ruleName]['{}positive'.format(key)] = list(
                                set(rule2all[ruleName]['{}positive'.format(key)] + positive))
                            rule2all[ruleName]['{}negative'.format(key)] = list(
                                set(rule2all[ruleName]['{}negative'.format(key)] + negative))
            for q in pquote:
                res = travelport(q, obj2port)
                if res:
                    innerport,innercom=res
                    rule2all[ruleName]['service']['port'] = list(set(rule2all[ruleName]['service']['port'] + innerport))
                    rule2all[ruleName]['service']['common'] = list(set(rule2all[ruleName]['service']['common'] + innercom))
                else:
                    rule2all[ruleName]['service']['common'].append(q)

        self.rule2all, self.obj2ipaddr, self.obj2port = rule2all, obj2ipaddr, obj2port

    def matchMultiPolicy(self, souall, desall, portall):
        """
        用处不大，待实现
        """
        pass

    def matchSinglePolicy(self, sou, des, port):
        """
        防火墙策略的匹配函数，输入源目端口返回相应的匹配信息
        :param sou:  网段三元组，不可为空
        :param des:  网段三元组，不可为空
        :param port: 端口三元组,可以为空
        :return: 返回matchs和isPolicyExist,前者为源目地址匹配到的策略详细，后者为是否连端口也完全匹配到
        """
        try:
            dll = CDLL('./subop.dll') #使用dll文件
        except:
            print("未能成功加载subop.dll，请检查是否有此文件或者此文件是否已被占用")
            return [], False
        dll.ifNetInTargetStruct.restype = POINTER(structToPython) #定义返回值
        dll.ifNetInTargetInner.restype= c_bool

        def negativeCal(negative, ipinfo):
            """
            内部函数，用于negative地址的匹配。 negative和positive是一对兄弟,nagative是exclude的地址，positive是正常的地址
            :param negative: exclude语句对应的地址会放到一条策略的x(s/d)negative里，这里的negative为网段存放网段的集合
            :param ipinfo: 想匹配的网络三元组
            :return: 匹配到的网络三元组
            """
            subs = []
            source=self.createStandardNet(ipinfo[0],ipinfo[1],ipinfo[2])
            for na in negative:
                it = getaddr(na)
                if it:
                    res=dll.ifNetInTargetInner(source,it)
                    if res:
                        subs.append(na)
                else:
                    print("[error] net {} can not be found in net2standard".format(na))
            return subs if subs else 1

        def getaddr(addr):
            """
            内部函数，用于网段映射到对应的标准网段
            :param addr: addr为一个网段字符串
            :return: addr对应的Standardnet
            """
            if addr in self.net2standard:
                return self.net2standard[addr]
            else:
                return None

        def matchSub(positive, ipinfo):
            """
            内部函数，用于匹配策略，会将参数传给dll，然后使用DLL中的函数进行计算
            此函数可以作为dll的调用示例。
            调用入口dll.ifNetInTargetStruct(source, tlist, length)中，source为标准网段，tlist为标准网段数组，length为所有标准网段的数量
            :param positive: 待匹配的网段列表
            :param ipinfo: 目标网络三元组
            :return: 匹配到则返回相应信息信息，未匹配到则返回None
            """
            source=self.createStandardNet(ipinfo[0],ipinfo[1],ipinfo[2])
            length = len(positive)
            if length==0:
                return ('0.0.0.0','0.0.0.0','c')
            tlist = (standardNet * length)()  #创建标准数组
            for i in range(length):
                t=getaddr(positive[i])  #获取对应的标准网段
                tlist[i] = t #填入标准网段数组
            length = c_int(length+1) #强制把python中的int类型转化成c中的int类型，防止出错
            res = dll.ifNetInTargetStruct(source, tlist, length)
            base = res.contents.base  # 返回的res是一个指针，从指针中取出对应的值
            offset = res.contents.offset
            itype = res.contents.itype
            dll.recycle_res()  # 回收内存，防止内存泄漏
            if itype == b'n':  # itype为'n',则未匹配到，返回None
                return None
            return (base,offset,itype)

        snegativeHit, dnegativeHit, matchs = [], [], []
        isPolicyExist = False
        for ruleName in self.rule2all:
            res = negativeCal(self.rule2all[ruleName]['snegative'], sou) #先匹配源的exclude
            if not res:
                continue  #匹配到则跳出，证明不在本策略范围内
            else:
                if isinstance(res, list):
                    snegativeHit = res
            smatch, dmatch = '', ''
            smatch = matchSub(self.rule2all[ruleName]['spositive'], sou) #匹配源地址
            if smatch: #如果匹配到源地址
                res = negativeCal(self.rule2all[ruleName]['dnegative'], des) #再匹配目的exclude
                if not res:
                    continue  #匹配到则跳出，证明不在本策略范围内
                else:
                    if isinstance(res, list):
                        dnegativeHit = res
                dmatch = matchSub(self.rule2all[ruleName]['dpositive'], des) #匹配目的地址

            if smatch and dmatch: #源目都匹配到，开始匹配端口
                mt = {'name': ruleName, 'action':self.rule2all[ruleName]['action'],'fromzone':self.rule2all[ruleName]['fromzone'],
                      'tozone':self.rule2all[ruleName]['tozone'],'sou': smatch, 'des': dmatch, "snegativeHit": snegativeHit,
                      "dnegativeHit": dnegativeHit, 'port': [], 'isFullMatch': False}
                t = self.rule2all[ruleName]['service']
                mt['port'] = mt['port'] + t['common'] +t['port']
                if port:
                    for p in t['port']:
                        pobject=self.port2standard[p]
                        if pobject.protocal==bytes(port[2],'ascii'):  #协议相同
                            if pobject.base<=port[0] and pobject.offset>=port[1]:
                                mt['isFullMatch'] = isPolicyExist = True
                                break
                    if port[2] in t['common']: #协议相同
                        mt['isFullMatch'] = isPolicyExist = True
                matchs.append(mt)
                if isPolicyExist==True:
                    return  matchs, isPolicyExist
        del dll #删除dll对象
        return matchs, isPolicyExist

    @property
    def infoSummary(self):
        """
        输出防火墙简略信息
        """
        msg = '该防火墙一共包含以下区域:\n'
        for zone in self.zone2interface:
            t = ','.join(self.zone2interface[zone])
            t = '无' if not t else t
            msg += "    {}({})\n".format(zone, t)
        msg += '该防火墙共有{}条rule,其中：\n'.format(len(self.rule2all))
        if not self.zonecount:
            self.zonecount = defaultdict(int)
            for rulename in self.rule2all:
                t = self.rule2all[rulename]
                fz = t['fromzone'] or 'unknown'
                tz = t['tozone'] or 'unknown'
                self.zonecount["{}->{}".format(fz, tz)] += 1
        for key in self.zonecount:
            msg += "    {} : {}\n".format(key, self.zonecount[key])
        a, p = len(self.obj2ipaddr), len(self.obj2port)
        msg += '该防火墙有{}个object-group定义，其中:\n'.format(a + p)
        aq, pq = len(self.addrQuote), len(self.portQuote)
        msg += '    地址定义{}条 ({}/{}已使用)\n    端口定义{}条 ({}/{}已使用)'.format(a, aq, a, p, pq, p)
        return msg

    def analyse(self, policy, isshort):
        """
        对传入的查询语句进行解析，然后传入matchSinglePolicy,并返回查询信息
        """
        if policy:
            addrs = addressMatch.findall(policy)
            if len(addrs) != 2:
                return "必须输入源目两个ip地址，遵循以下格式:\n   a.b.c.d/e\n   a.b.c.d-g.h.j.k\n   a.b.c.d g.h.j.k"
            else:
                sour, dest = addrs[0], addrs[1]
                sou = des = None
                port = ''
                for key, addr in zip(['s', 'd'], [sour, dest]):
                    sub, mask, flag = 0, 0, 'c'
                    if addr[5]:
                        sub = addr[5]
                        if addr[6]:
                            if int(addr[6]) == 32:
                                mask = '255.255.255.255'
                            elif int(addr[6])>0:
                                mask =self.num2ip(self.short2mask[addr[6]])
                            else:
                                mask="0.0.0.0"
                        else:
                            mask = '255.255.255.255'
                        flag = 'c'
                    elif addr[0]:
                        sub = addr[0]
                        mask =addr[2]
                        flag = 'w' if addr[1] else 'c'
                    elif addr[3]:
                        sub = addr[3]
                        mask = addr[4]
                        flag = 'r'
                    if key == 's':
                        sou = (sub, mask, flag)
                    else:
                        des = (sub, mask, flag)
                print("源地址为:{}\n目标地址为:{}".format(str(sou),str(des)))
                addrtmp = list(set(addrs[0])) + list(set(addrs[1]))
                t = re.split(r'[/ ",]', policy)[-1]
                portAna = re.compile(r'(?P<sympol>(?:lt|lte|gt|gte|eq|\<|\<=|\>|\>=|=))?'
                                     r'(?P<pro>tcp|udp|icmp)(?P<start>\d+)?(?:-(?P<end>\d+))?')
                if t not in addrtmp:
                    print("目标端口为:{}\n".format(t))
                    tp=portAna.match(t)
                    if tp:
                        dicttp=tp.groupdict()
                        nowbase=dicttp['start']
                        if dicttp['end']:
                            nowoffset=dicttp['end']
                        else:
                            nowoffset=0
                        if dicttp['sympol']:
                            bb,oo=self.portTrans(nowbase,nowoffset,dicttp['sympol'])
                        else:
                            if nowoffset:
                                bb,oo=self.portTrans(nowbase, nowoffset,"")
                            else:
                                bb,oo=self.portTrans(nowbase,nowoffset,'eq')
                        port=(bb,oo,dicttp['pro'])
                    else:
                        port=(0,0,t)
                else:
                    port=None
                matchs, isfullmatch = self.matchSinglePolicy(sou, des, port)
                if isshort:
                    return "【匹配到】" if isfullmatch else "未匹配到"
                else:
                    msg = '【匹配到】相应策略\n' if isfullmatch else "【未匹配到】相应策略\n"
                    if isfullmatch:
                        msg += "以下为匹配到策略的详细信息:\n"
                        for m in matchs:
                            if m['isFullMatch']:
                                msg += "->rule名:{}\n  区域:{}-->{}\n  action:{}\n  源地址:{}\n  目地址:{}\n  源exclude:{}\n  目exclude:{}\n  端口:{}\n".format(
                                m['name'], m['fromzone'] or "any",m['tozone'] or "any",m['action'],m['sou'], m['des'], m['snegativeHit'], m['dnegativeHit'],','.join(m['port'])
                                )
                        return msg
                    else:
                        if port:
                            if matchs:
                                msg += "但此不匹配为端口不匹配，可考虑在以下策略中增加{}端口:\n".format(port)
                                for m in matchs:
                                    msg += "->rule名:{}\n  区域:{}-->{}\n  action:{}\n  源地址:{}\n  目地址:{}\n  源exclude:{}\n  目exclude:{}\n  端口:{}\n".format(
                                        m['name'],m['fromzone'] or "any",m['tozone'] or "any",m['action'],m['sou'], m['des'], m['snegativeHit'], m['dnegativeHit'],(','.join(m['port']) or "any")
                                    )
                        else:
                            msg += "但源目地址已经匹配，以下为列出的端口:\n"
                            tin,yin = [],[]
                            for m in matchs:
                                msg += "->rule名:{}\n  区域:{}-->{}\n  action:{}\n  端口:{}\n".format(m['name'],m['fromzone'] or "any",m['tozone'] or "any", m['action'],(','.join(m['port'])) or "any")
                                tin += m['port']
                                if m['action']=='pass':
                                    if m['port']:
                                        tin += m['port']
                                    else:
                                        tin.append('any')
                                else:
                                    if m['port']:
                                        yin += m['port']
                                    else:
                                        yin.append('any')
                            tin,yin = list(set(tin)),list(set(yin))
                            msg += "所有pass端口:{}\n所有deny端口:{}".format(','.join(tin),','.join(yin))
                    return msg

    def exportExcel(self,name):
        """
        导出防火墙的详细信息到excel
        """
        try:
            import openpyxl
        except:
            print('无法导入openpyxl包,无法导出')
            return None

        def flat(mobj, objd, key=''):
            """
            用递归来扁平化字典
            """
            if isinstance(mobj, dict):
                for k in mobj:
                    newkey = key + '.' + k
                    flat(mobj[k], objd, newkey)
            elif isinstance(mobj, list) or isinstance(mobj, tuple) or isinstance(mobj, set):
                objd[key] = '\n'.join(mobj)
            elif isinstance(mobj, str):
                objd[key] = mobj
            else:
                return

        def flatten(mobj,getkey=False):
            """
            字典扁平化函数的入口
            """
            md = {}
            flat(mobj, md, '')
            if getkey:
                return [k for k in md]
            return md

        wb=openpyxl.Workbook()
        ws = wb.active
        ws.title = 'rule'
        ws1 = wb.create_sheet('address')
        ws2 = wb.create_sheet("port")
        flag=2
        rulec =[]
        for k in self.rule2all:
            rulec=flatten(self.rule2all[k],getkey=True)
            break
        if rulec:
            ws.cell(row=1, column=1, value="ruleName")
            for i in range(len(rulec)):
                ws.cell(row=1, column=i + 2, value=rulec[i])
            for rulename in self.rule2all:
                ws.cell(row=flag, column=1, value=rulename)
                res=flatten(self.rule2all[rulename])
                for j in range(len(rulec)):
                    ws.cell(row=flag,column=j+2,value=res[rulec[j]])
                flag+=1
        flag=2
        addrc=[]
        for k in self.obj2ipaddr:
            addrc=flatten(self.obj2ipaddr[k],getkey=True)
            break
        if addrc:
            ws1.cell(row=1, column=1, value="addrName")
            for i in range(len(addrc)):
                ws1.cell(row=1, column=i + 2, value=addrc[i])
            for obj in self.obj2ipaddr:
                ws1.cell(row=flag, column=1, value=obj)
                res=flatten(self.obj2ipaddr[obj])
                for j in range(len(addrc)):
                    ws1.cell(row=flag, column=j+2, value=res[addrc[j]])
                flag += 1
        flag=2
        portc = []
        for k in self.obj2port:
            portc = flatten(self.obj2port[k], getkey=True)
            break
        if portc:
            ws2.cell(row=1, column=1, value="portName")
            for i in range(len(portc)):
                ws2.cell(row=1, column=i + 2, value=portc[i])
            for obj in self.obj2port:
                ws2.cell(row=flag, column=1, value=obj)
                res = flatten(self.obj2port[obj])
                if res:
                    for j in range(len(portc)):
                        ws2.cell(row=flag, column=j+2, value=res[portc[j]])
                    flag += 1
        wb.save("{}.xlsx".format(name))

    def show(self,name,subname=""):
        """
        show命令用于打印信息
        """
        target=None
        if name=="rule" or name=="r":
            target=self.rule2all
        elif name=="address" or name=="a":
            target=self.obj2ipaddr
        elif name=="service" or name=="s":
            target=self.obj2port
        elif name=="standardnet" or name=="sn":
            target=self.net2standard
        elif name=="standardport" or name=="sp":
            target=self.port2standard
        if not subname:
            tmp=[key for key in target]
            l=len(tmp)
            if l%4==0:
                for i in range(int(l/4)):
                    print("{0:<31}{1:<31}{2:<31}{3:<31}\n".format(tmp[i*4],tmp[i*4+1],tmp[i*4+2],tmp[i*4+3]), chr(12288))
            else:
                for i in range(int(l/4)-1):
                    print("{0:<31}{1:<31}{2:<31}{3:<31}\n".format(tmp[i*4],tmp[i*4+1],tmp[i*4+2],tmp[i*4+3]), chr(12288))
                tmp=tmp[(l-l%4):]
                print("                 ".join(tmp))
        else:
            subname=subname.strip()
            if subname in target:
                try:
                    t=json.dumps(target[subname],indent=8)
                except:
                    if name == "standardnet" or name == "sn":
                        o=target[subname]
                        t="<standardnet object> base={}({}) offset={}({}) itype={}".format(self.num2ip(o.base),o.base,self.num2ip(o.offset),o.offset,str(o.itype))
                    elif name == "standardport" or name == "sp":
                        o=target[subname]
                        t="<standardpoet object> base={} offset={} protocal={}".format(o.base,o.offset,str(o.protocal))
                print("{}".format(t))
            else:
                print("无此对象")


cmdtransRaw="ls|dir\npwd|chdir\ntouch|cd .>\ncp|copy\nmv|move\nclear|cls\ndiff|comm|cmp|comp\n" \
            "rm|rm -f|del\ndate|time\ncat|type\nrm -rf|rm -r deltree\ntop|tasklist\nvi|notepad\n" \
            "env|set\nifconfig|ipconfig\nkill|taskkill /pid"
pathSep=re.compile(r'[/\\]')
platform=sys.platform

comTrans={} #翻译字典，下面这段用于填充内容到字典
for it in cmdtransRaw.split('\n'):
    items=it.split('|')
    for item in items:
        if item!=items[-1]:
            comTrans[item]=items[-1]



def outerSystemCommand(orinalpath):
    """
    代理执行dos命令或shell命令，仅支持路径的上下文保存
    """
    #orinalpath=orinalpath
    media="{}\\tmp.bin".format(orinalpath)
    fd=open(media,'w')
    fd.close()
    prepath = ''

    def systemCommandHandle(cmd):
        nonlocal prepath
        cs=cmd.strip().split()
        if cs and "win" in platform:  #当前windows系统，则对部分linux命令做翻译替换
            c1=comTrans.get(cs[0]) or cs[0]
            cmd="{} {}".format(c1,' '.join(cs[1:]))
        t="cd {} & {} & chdir>{}".format(prepath,cmd,media)
        os.system(t)  #执行命令
        with open(media,'r') as f: #读入新的路径
            prepath=f.read().strip('\n')
        return prepath

    def clean():
        os.remove(media)

    return systemCommandHandle,clean

def interactiveLoop():
    """
    实现简单的交互式命令行
    """
    cmd, currentDb,currentPath = '', 'noselect' ,os.getcwd()
    handle,clean=outerSystemCommand(currentPath)
    simplePath = currentPath.split("\\")[-1] if currentPath.split("\\") else ""
    dbs = {}

    def init(fullpath):
        """
        文件路径预处理，实例化H3CFwoperation对象，并读取配置
        """
        nonlocal currentDb
        if '.log' in fullpath or '.txt' in fullpath or '.' not in fullpath:
            try:
                fullpath = fullpath.replace('\\', '/')
                tmp = H3CFwoperation()
                tmp.configTrans(fullpath)
                fi = (fullpath.split('/')[-1]).split(".")[0]
                dbs[fi] = tmp
                currentDb = fi
                print("解析{}成功".format(fullpath))
            except:
                print("解析{}时出错，请检查文件格式".format(fullpath))
    """
    主循环，输入exit退出
    """
    while cmd != 'exit':
        cmd = input("[{} {}]>".format(currentDb,simplePath))
        cmd = cmd.strip()
        if cmd=="?" or cmd=="？":
            if not dbs:
                print("  init 用于初始化一个配置文件,后面跟一个log/txt结尾或无后缀的文件,如init h3c.log ,或者init *为读取当前文件夹下所有文本文件\n")
            else:
                print("  init          用于初始化一个配置文件,后面跟一个log/txt结尾或无后缀的文件,或者*为当前文件夹下所有文本文件\n"+
                      "  select        用于选择一个配置文件作为当前文件,查询操作均在当前配置文件上完成,使用\"select ?\"查看所有可用配置文件\n"+
                      "  export        导出一个配置文件为excel，默认为本程序所在目录,使用格式为:export 配置文件名(select ?查看可用配置文件名)\n"+
                      "  check         用于检查是否匹配rule,如需精简输出请使用check-s\n"+
                      "                格式为check 源地址,目标地址,端口(可缺省)，check关键字也可以省略,如:\n"+
                      "                25.48.64.64 wildcard 0.15.31.63,10.10.24.0 255.255.255.0,tcp7991\n"+
                      "                25.48.64.64~25.48.64.100,10.10.24.0 255.255.255.0,tcp7991\n"+
                      "  show          用于查看具体条目，show [rule(r)/address(a)/service(s)/standardnet(sn)/standardport(sp)] [具体名字](可缺省)\n"+
                      "  function      查看更多函数式或属性式命令\n"+
                      "  cmd           进入windows自带的命令行，进入后可以输入exit退回此命令行\n"+
                      "                (本命令行对linux的指令做了一部分的简单翻译，诸如ls之类的命令可以使用)"
                      ")"
                   )
            continue
        try:
            cmatch = commandMatch.match(cmd)
            if cmatch:
                cmddict = cmatch.groupdict()
                if cmddict['isinit']:
                    if cmddict['path']:
                        path = cmddict['path']
                        path= "\\".join([sp for sp in  pathSep.split(path) if sp.strip()])
                        if "\\" not in path:
                            path="{}\{}".format(currentPath,path)
                        if '*' in path:
                            path=path.replace("*",'.+')
                            for file in os.listdir(currentPath):
                                fullpath = "{}\{}".format(currentPath
                                                          .strip('/').strip('\\'), file)
                                if re.match(path.replace("\\","\\\\"),fullpath):
                                    if os.path.isfile(fullpath):
                                        init(fullpath)
                        elif os.path.isfile(path):
                                init(path)
                        else:
                            try:
                                for file in os.listdir(path):
                                    if os.path.isfile(file):
                                        if '.log' in file or '.txt' in file or '.' not in file:
                                            print(file)
                            except:
                                msg = "                * : all file\n xx.log/xx.txt/xx : single file"
                                print(msg)
                    else:
                        basepath = os.getcwd()
                        for file in os.listdir(basepath):
                            if os.path.isfile(file):
                                if '.log' in file or '.txt' in file or '.' not in file:
                                    print(file)
                elif cmddict['alldb']:
                    print(' '.join([i for i in dbs]))
                elif cmddict['db']:
                    if cmddict['db'] in dbs:
                        currentDb = cmddict['db']
                        print('ok')
                    else:
                        print('{} does not exist'.format(cmddict['db']))
                elif cmddict['spolicy'] or cmddict['policy']:
                    isshort = False
                    policy = ""
                    if cmddict['spolicy']:
                        t = addressMatch.match(cmddict['spolicy'])
                        policy = cmddict['spolicy']
                    else:
                        t = addressMatch.match(cmddict['policy'])
                        policy = cmddict['policy']
                        if cmddict['isshort']:
                            isshort = True
                    if t:
                        if currentDb != 'noselect':
                            res = dbs[currentDb].analyse(policy, isshort)
                            print(res)
                        else:
                            print("请选择一个配置后再进行查询")
                    else:
                        if policy.strip() in dbs:
                            print(dbs[cmddict['spolicy'].strip()].infoSummary())
                            continue
                        peaches=postcommandMatch.findall(policy)
                        if peaches:
                            peach=peaches[0]
                            if peach[0]:
                                targetdb=peach[0]
                                if targetdb in dbs:
                                    dbs[targetdb].exportExcel(targetdb)
                                    print('导出完成')
                                else:
                                    print("无此配置,请从以下项导出:\n {}".format(','.join([i for i in dbs])))
                            elif peach[1]:
                                if currentDb == 'noselect':
                                    print('请先init一个log，有配置文件才能使用show命令')
                                    continue
                                if peach[2]:
                                    dbs[currentDb].show(peach[1],peach[2])
                                else:
                                    dbs[currentDb].show(peach[1])
                        else:
                            try:
                                cccmd="print(dbs[currentDb].{})".format(policy.strip())
                                exec(cccmd)
                            except:
                                #handle(cmd)
                                for word in "hello0thank0you0yep0nope0music0dj0sing0wow0damn0fuck0sex0shit0hc3".split('0'):
                                    if cmd.__contains__(word):
                                        for s, st in zip("hello?|thank|you|are|you|ok?".split("|"),
                                                         [0.8, 0.5, 0.8, 0.4, 0.4, 3]): print(
                                            "{0:<10}".format(s) * 6);time.sleep(st)

                                try:
                                    prepath=handle(cmd)
                                    currentPath=prepath
                                    simplePath=prepath.split("\\")[-1] if prepath.split("\\") else ""
                                except:
                                    print("命令执行错误")

        except Exception as e:
            print("执行命令过程中发生错误:{}".format(e))

if __name__ == "__main__":
    interactiveLoop()
    pass

"""
全匹配组测试
import time
a = H3CFwoperation()
a.configTrans("D:\DevProj\Djproj\\feemanager\h3c.log")
pcount=0
for k in a.rule2all:
    t=a.rule2all[k]
    s,d,p=t['spositive'],t['dpositive'],t['service']['port']
    for ss in s:
        for dd in d:
            for pp in p:
                command="{},{},{}".format(ss,dd,pp)
                #print(command)
                res=a.analyse(command,isshort=True)
                pcount+=1
                if res!="【匹配到】":
                    print(command)
"""